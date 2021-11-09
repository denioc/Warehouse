"""
Модуль работы с файлами Excel
При записи данных в таблицу следует учитывать, что в Excel нумерация ячеек начинается с 1, а не с 0.
При создании экземпляра класса программа запускает приложение Excel.Aplication. Если создается еще один 
экземпляр класса или в ОС уже открыта таблица Excel - приложение Excel будет запускаться повторно, но
на работу это ни как не влияет. Реализовать проверку, запущено ли приложение, не удалось.
Необходимо проработать вопрос закрытия приложения Excel.Application:
- если после работы с файлом его закрывать, то будут закрываться все открытые файлы Excel (не точно);
- если не закрывать приложение совсем, то неизвестно дальнейшее поведение Excel в среде ОС -
(возможно переполнение памяти, открыто слишком много Excel приложений, или просто бардак в службах ОС).
Пока решено приложение Excel не закрывать.

pip install win32com

pip install openpyxl
"""
# Модуль win32com.client позволяет использовать COM объекты Windows для использования Microsoft Excel и Word.
# Это позволит выполнять с файлом все операции, которые позволяет делать обычный Excel с использованием VBA.
# Однако модуль работает только в среде Windows, поскольку использует стандартные средства Windows win32com
#import win32com.client

# Популярный модуль работы с Excel. Позволяет работать с таблицами Excel "на лету"
from openpyxl import Workbook		# Для создания новой таблицы
from openpyxl import load_workbook	# Для чтения талицы
from kivy.metrics import mm
#from kivy.utils import rgba
from kivy.utils import get_color_from_hex

XL_FILE = u'Общий.xlsx'
XL_SHEET = 'SPR4-03'

# Класс работает на основе модуля win32com. Не подходит для кроссплатформенности
class winExcel:
    #Excel = win32com.client.Dispatch("Excel.Application")
    #File = u'D:\\LMT\\SCLAD\\Складской учет\\Общий.xlsx']
    File = u'Общий.xlsx'
    Sheet = 'SPR4-03'

    def __init__(self, **kwargs):
        """ Запуск приложения Excel"""
        self.Excel = win32com.client.Dispatch("Excel.Application") #Если приложение не запущено

    def destroy(self):
        """ Закрыть приложение Excell"""
        self.WorkBook.Close()
        self.Excel.Quit()

    def open_file(self):
        """ Открыть файл Excel для работы """
        self.WorkBook = self.Excel.Workbooks.Open(self.File)

    def open_sheet(self):
        """ Открыть лист Excel для работы """
        self.Sheet = self.WorkBook.Worksheets(self.Sheet)

    def close(self, save=True):
        if save:
            self.WorkBook.Save()
            print(f"XLS file saved!")
        self.WorkBook.Close()
        self.Excel.Quit()

    def read(self):
        """ Чтение таблицы """
        OrderCode = [r[0].value for r in self.Sheet.Range("F4:F32")]       # OrderCode
        TotalCount = [r[0].value for r in self.Sheet.Range("I4:I32")]       # количество
        Analog = [r[0].value for r in self.Sheet.Range("G4:G32")]       # Analog
        print(f"Данные с таблицы прочитаны: \n{OrderCode , TotalCount, Analog}")
        return {'OrderCode':OrderCode, 'TotalCount':TotalCount, 'Analog':Analog}

    def write_single(self, sheet, value):
        """ Запись одной ячейки в таблицу """
        sheet.Cells(9,1).value = value

    def write_range(self, sheet_result, row=1, col=1, data=[], wr_type='column'):
        """
        Запись диапазона ячеек в таблицу
        row: номер первой строки для записи
        col: номер первого столбца для записи
        data: массив данных для записи
        wr_type: тип записи. 'column' - запись в столбец; 'row' - запись в строку
        """
        if wr_type not in {'column', 'row'}: 
            print(f"[ОШИБКА]: Ошибка настроек записи таблицы")
            return
        for item in data:
            sheet_result.Cells(row, col).value = item
            if wr_type == 'column': row += 1
            if wr_type == 'row': col += 1

class Excel:
	# wb = Workbook() # Создать новую книгу
	#FILE = u'D:\\LMT\\SCLAD\\Складской учет\\Общий.xlsx'
	"""
	def __init__(self, file):
		# Занести сюда загрузку книги??? load_workbook
		print("Excell init")
		print(dir(self))
		#self.FILE = file
		self.wb=load_workbook(file)
		#try:
			#self=load_workbook(file)
		#except:
			#print("Ошибка загрузки книги Excel! Проверьте имя и путь к файлу.")
		#print(f" SHEET_NAMES: {self.sheetnames}")
		#print(self.worksheets)
			#return False
		#return wb
	"""
	def open_book(file, data_only=True):
		""" Функция открывает существующую книгу Excel.
		  Флаг data_only определяет чтение формул: 
		  True - читаются пересчитанные в последний раз по формулам данные; 
		  False - читаются непосредственно сами формулы """
		try:
			wb = load_workbook(file, data_only=data_only)
		except:
			print("ERROR: Ошибка открытия файла!!!")
			wb = None
		finally:
			return wb

	def read_sheet(workbook, sheet):
		""" Функция читает и возвращает лист в книге Excell"""
		try:
			sh = workbook[sheet]
		except:
			print(f"ERROR: Ошибка чтения страницы {sheet} в документе {workbook.__name__}")
			sh = None
		finally:
			return sh
			
	def get_cols_rows_names(range_cells):
		""" Функция возвращает 2 списка: с именами столбцов и строк.
		Входящий range_cells д.б. типа string и может иметь формат: AB4:BZ34, DC8:CF8"""
		def str_to_alpha_digit(str):
			alpha = ""
			digit = ""
			for ch in str:
				if ch.isalpha():
					alpha += ch
				elif ch.isdigit():
					digit += ch
			return alpha, int(digit)
			
		if not isinstance(range_cells, str):
			return None
		range_cells = range_cells.replace(' ', '') #удалили все пробелы
		ranges = range_cells.split(',')
		col_names = []
		row_names = []
		for rng in ranges:
			rng = rng.split(':')
			start_alpha, start_digit = str_to_alpha_digit(rng[0])
			stop_alpha, stop_digit = str_to_alpha_digit(rng[1])
			col_names += [s for s in Excel._char_range(start_alpha, stop_alpha)]
			row_names += [n for n in Excel._num_range(start_digit, stop_digit)]
		return col_names, row_names

	def _ord_col_name(str):
		""" Функция переводит строчное имя столбца в число """
		num = 0
		for i, ch in enumerate(str[::-1]):
			n = ord(ch)-64
			num += n*(26**i)
		return num
		
	def _chr_col_name(num):
		""" Функция переводит число в строчное имя столбца """
		str=""
		i = 1
		j = 0
		k = 0
		while num>k:
			m = 26**i
			n = num%m
			k = 26**j
			nn = n//k
			if nn == 0:
				nn = 26
			num -= nn*k
			i += 1
			j += 1
			str = chr(nn+64) + str
		return str

	def _char_range(start, stop, step=1):
		""" Функция-генератор аналогично range (возвращает обозначения столбцов в диапазоне). 
		Вернет буквы в обратном порядке, если так указаны start, stop """
		nst = Excel._ord_col_name(start)
		nsp = Excel._ord_col_name(stop)
		if nst > nsp and step > 0:
			step *= -1
		for num in range(nst, nsp+step, step):
			col_name = Excel._chr_col_name(num)
			yield col_name
		
	def _num_range(start, stop, step=1):
		""" Функция-генератор аналогично range (возвращает числа в диапазоне). 
		Вернет числа в обратном порядке, если так указаны start, stop """
		if start > stop and step > 0:
			step *= -1
		for num in range(start, stop+step, step):
			yield num
			
	COLORS = {'00000000': [1,1,1,1]} # Словарь хранит сконвертированные значения цветов {argb: rgba}.
	
	def _argb_to_rgba(argb):
			# Функция переводит цвет формата argb в формат rgba.
			# Чтобы не конвертировать цвет каждый раз используется словарь COLORS с сохраненными отконверитрованными значениями
			if argb in Excel.COLORS:
				color = Excel.COLORS[argb]
			else:
				color = get_color_from_hex(argb[2:] + argb[:2])
				Excel.COLORS.update({argb: color})
			return color
			
	def view_sheet(sheet, widget):
		# Функция выводит на виджет страницу из таблицы Excel
		# sheet - страница таблицы
		# widget - виджет вывода таблицы из модуля table.py (класс WorkTable)
		#head_color = (0.8,0.8,0.8,1)	# Цвет фона заголовка таблицы
		#table_color = (0.8,0.98,0.98,1)
		#print("??????", table_color)
		# Выгрузить из таблицы все данные. Диапазон будет соответствовать размеру таблицы
		table_data = []
		# ws.values - генератор, возвращает все значения в таблице
		# Читаем значения из таблицы
		for row in sheet[sheet.dimensions]: #sheet.values:
			for cell in row:
				data = str(cell.value) if cell.value!=None else ""
				color = Excel._argb_to_rgba(cell.fill.fgColor.rgb)
				table_data.append({'text': data, 'background_color': color, 'focus': False})
		# Читаем размеры ячеек
		cd = {Excel._ord_col_name(k[0])-1: mm(k[1].width) for k in sheet.column_dimensions.items()}
		widget.columns_width = cd
		# Читаем цвета ячеек
		#cc = sheet['A12'].fill.fgColor.rgb
		#print(dir(sheet.dimensions))
		#print(sheet.print_title_cols)
		#widget.headers = [{'text': f"head {c}", 'background_color': head_color, 'focus': False} for c in range(len(row))]
		widget.cols = len(row)
		widget.data = table_data
		#widget.data = [{'text': str(data) if data!=None else "", 'background_color': table_color, 'focus': False} for data in table_data]
		#widget.data = [{'text': str(data) if data!=None else ""} for data in table_data] # без задания цвета
		
	def view_range_cells(sheet, data_range, widget):
		# Функция читает и выводит в виджет кусок таблицы с вычитанной шириной столбцов. Высота строк пока не применяется
		# sheet - страница таблицы
		# data_range - диапазон ячеек для чтения и вывода. Д.б. строка в формате 'F3:AA12'
		# widget - виджет вывода таблицы из модуля table.py (класс WorkTable)
		#headers_range = 'F3:J3'
		#data_range = 'F4:J32'
		#head_color = (0.8,0.8,0.8,1)	# Цвет фона заголовка таблицы
		table_color = (0.8,0.98,0.98,1)
		col_range, row_range = Excel.get_cols_rows_names(data_range)
		cols_width = {pos: mm(sheet.column_dimensions[col].width) for pos, col in enumerate(col_range)}
		rows_height = {pos: mm(sheet.row_dimensions[row].height) for pos, row in enumerate(row_range)}
		widget.columns_width = cols_width
		# Записываем данные в таблицу
		#head_data = [c.value for r in sheet[headers_range] for c in r]
		table_data = [c.value for r in sheet[data_range] for c in r]
		widget.data = [{'text': str(data) if data!=None else "", 'background_color': table_color, 'focus': False} for data in table_data]


if __name__ == '__main__':
	from kivymd.app import MDApp
	from kivy.uix.boxlayout import BoxLayout
	from kivy.uix.textinput import TextInput
	from kivy.uix.button import Button
	from kivy.metrics import mm
	class TestApp(MDApp):
		def build(self):
			bl = BoxLayout(orientation = 'vertical')
			bl.size = (800, 600)
			self.ti = TextInput()
			b = Button(text= "Test Button", on_press=self.press_button)
			bl.add_widget(self.ti)
			bl.add_widget(b)
			return bl

		def press_button(self, instance):
			print("Test button pressed")
			xl = Excel
			wb = xl.open_book(XL_FILE)
			#ws = xl.read_sheet(wb, XL_SHEET)
			ws = wb[XL_SHEET] #аналог read_sheet
			# Получить размеры ячейки:
			print(f"row = {ws.row_dimensions[3].height}")
			print(f"col = {ws.column_dimensions['A'].width}")
			# Получить названия строк диапазона
			#print(f"==={ws['A4:F32'].column_dimensions}")
			# Получить  размеры ячеек таблицы
			data_range = 'B4:F32, AZ10:AB33'
			cr, rr = xl.get_cols_rows_names(data_range)
			#print(f"range col = {cr}")
			#print(f"range row = {rr}")
			# Ширина столбцов. Имена столбцов числового типа (нумерация столбцов с нуля)
			cw = {pos: mm(ws.column_dimensions[col].width) for pos, col in enumerate(cr)}
			# Ширина столбцов. Имена столбцов числового типа (оригинальная нумерация столбцов)
			cw2 = {Excel._ord_col_name(col): mm(ws.column_dimensions[col].width) for  col in cr}
			# Ширина столбцов. Имена столбцов строчного типа
			cw3 = {col: mm(ws.column_dimensions[col].width) for col in cr}
			# Высота строк (нумерация строк с нуля):
			rh = {pos: ws.row_dimensions[row].height for pos, row in enumerate(rr)}
			# Высота строк (оригинальная нумерация строк):
			rh2 = {row: ws.row_dimensions[row].height for row in rr}
			print(f"col_width = {cw}")
			print(f"row_height = {rh}")
			
			# Выгрузить из таблицы все данные. Диапазон будет соответствовать размеру таблицы
			#for row in ws.values:
			#	for cell in row:
			#		print(cell)
			# Прочитать построчно таблицу
			#for row in ws.rows:
				#for cell in row:
				#	print(cell.value)
			#print(ws[ws.dimensions]) # Возвращает всю таблицу - кортеж строк, каждая строка - кортеж объектов ячеек в строке
			
			
			# Вывод данных в textinput
			print(dir(ws))
			print(ws['F5'].value)
			self.ti.text = str(wb.sheetnames)
			self.ti.text += "\n\n"
			self.ti.text += str(ws['F4'].value)
			self.ti.text += "\n\n"
			OrderCode = [r[0].value for r in ws['F4:F32']]		# Значения диапазона ячеек
			data = [c.value for r in ws['F4:J32'] for c in r]
			print(data)
			self.ti.text += str(OrderCode)


	TestApp().run()