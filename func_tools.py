"""
Модуль включает функции для выполнения бизнес-логики программы
"""
# pip install openpyxl

#import numpy
from openpyxl import Workbook, workbook		# Для создания новой таблицы
from openpyxl import load_workbook	# Для чтения талицы
from openpyxl.styles import Font, Alignment, PatternFill, Color
from kivy.metrics import mm, cm, dp, inch, pt, sp
from kivy import resources
from kivy.config import Config
#from openpyxl.utils.units import *
from openpyxl.utils.cell import column_index_from_string
#from kivy.utils import rgba
from kivy.utils import get_color_from_hex
from openpyxl.styles.colors import ColorList
from openpyxl.xml.constants import *
import pandas
from pandas.core.frame import DataFrame
from openpyxl.utils.dataframe import dataframe_to_rows


"""
sheet.row_dimensions.items() получает словарь со строками, где размеры и цвета отличаются от стандартных (где либо
цвет, либо размер (высота) изменен).
sheet.row_dimensions.clear - очищает размеры и цветовые выделения строк, но если строки еще есть - 
то в непустых ячейках цвет сохраняется и размер тоже, если он больше сандартного. 
При этом (после очистки) sheet.row_dimensions.item остается пустой даже после команд row_dimensions.update()
и sheet.calculate_dimensions().
Получается таблица ровно такая же, как мы ее читаем и выводим в виджет Table (цвета помечаются не по всей строке,
а по ячейкам с данными).
Видимо, помимо атрибута fgColor есть еще где-то значения цветов
Команды для информации:
print("max row:", sheet.max_row)
print("count dimensions", len(sheet.row_dimensions.items()))
print(sheet.row_dimensions.items())
print(dir(sheet.row_dimensions))
# delete_rows(start_row, count_rows)
sheet.delete_rows(1, sheet.max_row)	# Удаляет все строки, форматиррвание остается
#sheet.reset_dimensions() # Почему нет в классе, хотя в документации есть такая функция (возможно она только в режиме для чтения)
sheet.row_dimensions.clear()
#sheet.calculate_dimension()
#sheet.row_dimensions.update()	# ничего не происходит
#sheet.column_dimensions.setdefault()	# нужен какой-то аргумент
"""

"========================================================================================"
"""
Набор функций для конвертации цвета типа 'theme' (параметры theme и tint) в цвет типа rgb.
Источник: 
https://stackoverflow.com/questions/58429823/getting-excel-cell-background-themed-color-as-hex-with-openpyxl

Пример использования решения:
wb = load_workbook(filename, data_only=True)
theme = cell.fill.start_color.theme	# *либо cell.fill.fgColor.theme
tint = cell.fill.start_color.tint # *либо cell.fill.fgColor.tint
color = theme_and_tint_to_rgb(wb, theme, tint)
*Примечание: print(dir(cell.fill)) не отображает атрибута start_color, есть fgColor, но при этом все работает

"""
from colorsys import rgb_to_hls, hls_to_rgb
# From: https://stackoverflow.com/questions/58429823/getting-excel-cell-background-themed-color-as-hex-with-openpyxl/58443509#58443509
#   which refers to: https://pastebin.com/B2nGEGX2 (October 2020)
#       Updated to use list(elem) instead of the deprecated elem.getchildren() method
#       which has now been removed completely from Python 3.9 onwards.
#

#https://bitbucket.org/openpyxl/openpyxl/issues/987/add-utility-functions-for-colors-to-help

RGBMAX = 0xff  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:
# https://social.msdn.microsoft.com/Forums/en-US/e9d8c136-6d62-4098-9b1b-dac786149f43/excel-color-tint-algorithm-incorrect?forum=os_binaryfile#d3c2ac95-52e0-476b-86f1-e2a697f24969

def rgb_to_ms_hls(red, green=None, blue=None):
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))

def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    if lightness is None:
        hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)

def rgb_to_hex(red, green=None, blue=None):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    if green is None:
        red, green, blue = red
    return ('%02x%02x%02x' % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()


def get_theme_colors(wb):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]

    colors = []

    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        for i in list(accent): # walk all child nodes, rather than assuming [0]
            if 'window' in i.attrib['val']:
                colors.append(i.attrib['lastClr'])
            else:
                colors.append(i.attrib['val'])

    return colors

def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))

def theme_and_tint_to_rgb(wb, theme, tint):
    """Given a workbook, a theme number and a tint return a hex based rgb"""
    rgb = get_theme_colors(wb)[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))

"========================================================================================"

FONT_SUBSTITUTION = {'Courier New': 'Cour', 'Times New Roman':'Times', 'Consolas': 'Consola'}
# Приписка стиля вконце имени шрифта, key = (Bold, Italic), boolean type.
FONT_STYLE_ENDINDEX = {(0, 0): '', (1, 0): ['b', 'bd', '-Bold', 'Bold'], # Bold
	(0, 1): ['i', '-Italic', 'Italic'], # Italic
	(1, 1): ['bi', 'z', '-BoldItalic', 'BoldItalic']} # Bold+Italic

class Excel:
	# wb = Workbook() # Создать новую книгу
	#FILE = u'D:\\LMT\\SCLAD\\Складской учет\\Общий.xlsx'

	def open_book(file, data_only=True):
		""" Функция открывает существующую книгу Excel.
		Флаг data_only определяет чтение формул: 
		True - читаются пересчитанные в последний раз по формулам данные; 
		False - читаются непосредственно сами формулы
		Функция возвращает прочитанную книгу
		"""
		try:
			wb = load_workbook(file, data_only=data_only)
		except:
			print("ERROR: Ошибка открытия файла!!!")
			wb = None
		finally:
			return wb

	def read_sheet(workbook, sheet):
		""" Функция читает и возвращает лист в книге Excell"""
		try:
			sh = workbook[sheet]
		except:
			print(f"ERROR: Ошибка чтения страницы {sheet} в документе {workbook.__name__}")
			sh = None
		finally:
			return sh

	def save_book(workbook, filename):
		""" Функция сохраняет существующую книгу Excel."""
		try:
			workbook.save(filename)
			print(f"Данные сохранены в {filename}")
		except:
			print("""ERROR: Не удается сохранить файл!!! 
			Возможно, он используется другим приложением. Закройте файл и попробуйте снова.""")


	def get_cols_rows_names(range_cells):
		""" Функция возвращает 2 списка: с именами столбцов и строк.
		Входящий range_cells д.б. типа string и может иметь формат: AB4:BZ34, DC8:CF8"""
		def str_to_alpha_digit(str):
			alpha = ""
			digit = ""
			for ch in str:
				if ch.isalpha():
					alpha += ch
				elif ch.isdigit():
					digit += ch
			return alpha, int(digit)
			
		if not isinstance(range_cells, str):
			return None
		range_cells = range_cells.replace(' ', '') #удалили все пробелы
		ranges = range_cells.split(',')
		col_names = []
		row_names = []
		for rng in ranges:
			rng = rng.split(':')
			start_alpha, start_digit = str_to_alpha_digit(rng[0])
			stop_alpha, stop_digit = str_to_alpha_digit(rng[1])
			col_names += [s for s in Excel._char_range(start_alpha, stop_alpha)]
			row_names += [n for n in Excel._num_range(start_digit, stop_digit)]
		return col_names, row_names

	def _ord_col_name(str):
		""" Функция переводит строчное имя столбца в число """
		num = 0
		for i, ch in enumerate(str[::-1]):
			n = ord(ch)-64
			num += n*(26**i)
		return num
		
	def _chr_col_name(num):
		""" Функция переводит число в строчное имя столбца """
		str=""
		i = 1
		j = 0
		k = 0
		while num>k:
			m = 26**i
			n = num%m
			k = 26**j
			nn = n//k
			if nn == 0:
				nn = 26
			num -= nn*k
			i += 1
			j += 1
			str = chr(nn+64) + str
		return str

	def _char_range(start, stop, step=1):
		""" Функция-генератор аналогично range (возвращает обозначения столбцов в диапазоне). 
		Вернет буквы в обратном порядке, если так указаны start, stop """
		nst = Excel._ord_col_name(start)
		nsp = Excel._ord_col_name(stop)
		if nst > nsp and step > 0:
			step *= -1
		for num in range(nst, nsp+step, step):
			col_name = Excel._chr_col_name(num)
			yield col_name
		
	def _num_range(start, stop, step=1):
		""" Функция-генератор аналогично range (возвращает числа в диапазоне). 
		Вернет числа в обратном порядке, если так указаны start, stop """
		if start > stop and step > 0:
			step *= -1
		for num in range(start, stop+step, step):
			yield num
			
	def clear_sheet(sheet, del_formatting=True):
		# Очищает весь лист построчно вместе с размерами и цветами строк. 
		# Настройки столбцов не меняются
		# del_formatting: True - очищается форматирование; False - форматирование остается
		# delete_rows(start_row, count_rows)
		sheet.delete_rows(1, sheet.max_row)	# Удаляет все строки, форматиррвание остается
		if del_formatting:
			sheet.row_dimensions.clear()	# удаляет настройки строк

	def create_template(sheet):
		# Функция создает шаблон таблицы результата
		r = sheet.max_row + 2
		sheet.cell(row=r, column=1).value = "OrderCode"
		sheet.cell(row=r, column=2).value = "Analog"
		sheet.cell(row=r, column=3).value = "Необходимо"
		sheet.cell(row=r, column=4).value = "Насчитано"
		sheet.cell(row=r, column=5).value = "Позиция"
		sheet.cell(row=r, column=6).value = "Закуплено"
		sheet.cell(row=r, column=7).value = "Израсходованно"
		sheet.cell(row=r, column=8).value = "Остаток"
		sheet.cell(row=r, column=9).value = ""
		sheet.cell(row=r, column=10).value = "В закупку"
		grey_fgcolor = Color(theme=0,  tint=-0.1499984740745262)
		for i in range(10):
			sheet.cell(r, i+1).font = Font(bold=True)	# Жирный шрифт
			sheet.cell(r, i+1).alignment = Alignment(horizontal="center", vertical="center")	# выравнивание
			sheet.cell(r, i+1).fill = PatternFill("solid", fgColor=grey_fgcolor)	# Заливка
			#sheet.cell(r, 1).fill = PatternFill("solid") #, fgColor="DDDDDD")


	COLORS_rgb = {'00000000': [1,1,1,1]} # Словарь хранит сконвертированные значения цветов {argb: rgba} (тип rgb).
	COLORS_theme = {} # Словарь хранит сконвертированные значения цветов {theme&tint: rgba} (тип theme).
	
	def _argb_to_rgba(argb):
		# Функция переводит цвет формата argb в формат rgba.
		# Чтобы не конвертировать цвет каждый раз используется словарь COLORS 
		# с сохраненными отконвертированными значениями
		if argb in Excel.COLORS_rgb:
			color = Excel.COLORS_rgb[argb]
		else:
			color = get_color_from_hex(argb[2:] + argb[:2])
			Excel.COLORS_rgb.update({argb: color})
		return color

	def _theme_to_rgba(workbook, theme, tint):
		# Конвертируем theme&tint в rgba
		key = str(theme) + ';' + str(tint)
		if key in Excel.COLORS_theme:
			color = Excel.COLORS_theme[key]
		else:
			color_hex = theme_and_tint_to_rgb(workbook, theme, tint) + 'FF'
			color = get_color_from_hex(color_hex)
			Excel.COLORS_theme.update({key: color})
		return color
		
	def get_fill_color(cell):
		# Функция возвращает цвет ячейки в формате rgba
		color = [1,1,1,1] # default value
		if cell.fill.fgColor.type == 'rgb':
			color = Excel._argb_to_rgba(cell.fill.fgColor.rgb)
			#color = Excel._argb_to_rgba(cell.fill.start_color.rgb)
		elif cell.fill.fgColor.type == 'theme':
			# Получаем номер theme и значение tint
			theme = cell.fill.fgColor.theme
			tint = cell.fill.fgColor.tint
			# Конвертируем theme&tint в rgba
			color = Excel._theme_to_rgba(cell.parent.parent, theme, tint)
		elif cell.fill.fgColor.type == 'indexed':
			color = [1,1,1,1]	# Сделать функцию пересчета параметров index -> rgba 
		return color

	def get_font(obj):
		# Функция возвращает имя и размер шрифта для столбца, строки, или ячейки
		# Для некоторых ячеек шрифт может быть не указан, поэтому будет возвращен шрифт по-умолчанию. Нет, шрифт надо возвращать какой есть, а в таблице делать обработку
		# В Excel размер шрифта представляется в пунктах (пт). pixels = 11(пт)*96(dpi)/72
		ofn = obj.font.name #if obj.font.name != None #else 'Roboto'
		fn = FONT_SUBSTITUTION.get(ofn, ofn)
		fs = round(obj.font.sz*96/72, 1) if obj.font.sz != None else None #15
		#print(obj.font.sz)
		return fn, fs

	def find_font_style(obj):
		# Функция ищет шрифты нужного стиля (жирный, курсив, жирный курсив).
		# В файле шрифта со стилем после имени шрифта обычно добавляется приписка с обозначением стиля.
		# Бывают разные варианты приписок, функция ищет среди часто используемых.
		# === Можно ввести таблицу замены имен шрифтов, т.к. у некоторых имя шрифта не совпадает с именем файла (напр. Courer New= courer.ttf, Times New Roman = times.ttf). Так же это позволит ставить аналоги шрифтов
		# Если нужный стиль найден, возвращает имя этого шрифта.
		# Если стиль не найден, возвращается '"".
		fn = obj.font.name
		font_name = FONT_SUBSTITUTION.get(fn, fn)
		b = obj.font.bold
		i = obj.font.italic
		if not resources.resource_find(font_name + '.ttf'):
			font_name = 'Roboto'
		style = FONT_STYLE_ENDINDEX.get((b, i))
		for stl in style:
			fn = font_name + stl + '.ttf'
			if resources.resource_find(fn):
				return fn[:-4]	# Можно возвращать и с .ttf, но так красивее
		return ''

	def get_alignment(obj):
		# Функция возвращает горизонтальное выравнивание текста
		halign = obj.alignment.horizontal
		if halign == None:
			halign = 'auto'
		#valign = obj.alignment.vertical
		return halign

	def get_column_width(sheet):
		# Функция возвращает словарь с номером и шириной столбца {0: width}. Номера начинаются с 0.
		width = {}
		for col in sheet.column_dimensions.items():
			if col[1].width:
				#width.update({Excel._ord_col_name(col[0])-1: col[1].width})
				width.update({column_index_from_string(col[0])-1: col[1].width})
		return width

	def get_row_height(sheet):
		# Функция возвращает словарь с номером и высотой строки {0: width}. Номера начинаются с 0.
		height = {}
		for row in sheet.row_dimensions.items():
			if row[1].height:
				height.update({row[0]-1: row[1].height})
		return height

	def view_sheet(sheet, widget):
		# Функция выводит на виджет страницу из таблицы Excel
		# sheet - страница таблицы
		# widget - виджет вывода таблицы из модуля table.py (класс WorkTable)
		# Настраиваем шрифт
		# Сначала применяем вычитанный шрифт первого столбца. для всей таблицы, 
		# затем локально применяем к ячейкам настройки шрифта (bold, italic)
		font_name, font_size = Excel.get_font(sheet.column_dimensions['A'])
		widget.set_font(font_name, font_size)
		#print(font_name, font_size)
		# Выгрузить из таблицы все данные. Диапазон будет соответствовать размеру таблицы
		table_data = []
		# Читаем значения из таблицы
		for row in sheet[sheet.dimensions]: #sheet.values:
			for cell in row:
				data = str(cell.value) if cell.value!=None else ""
				color = Excel.get_fill_color(cell) # Читаем цвет ячейки
				halign = Excel.get_alignment(cell)
				table_data.append({'text': data, 'background_color': color, 'halign': halign})
				if cell.font.bold or cell.font.italic:
					font_name = Excel.find_font_style(cell)
					if font_name:
						table_data[-1].update({'font_name': font_name})
		# Настраиваем размеры ячеек
		dim = Excel.get_column_width(sheet)
		widget.columns_width = dim
		dim = Excel.get_row_height(sheet)
		widget.rows_height = dim
		
		widget.cols = len(row)
		widget.data = table_data
		#widget.data = [{'text': str(data) if data!=None else "", 'background_color': table_color, 'focus': False} for data in table_data]
		#widget.data = [{'text': str(data) if data!=None else ""} for data in table_data] # без задания цвета
		
	def view_range_cells(sheet, data_range, widget):
		# Функция читает и выводит в виджет кусок таблицы с вычитанной шириной столбцов. Высота строк пока не применяется
		# sheet - страница таблицы
		# data_range - диапазон ячеек для чтения и вывода. Д.б. строка в формате 'F3:AA12'
		# widget - виджет вывода таблицы из модуля table.py (класс WorkTable)
		#headers_range = 'F3:J3'
		#data_range = 'F4:J32'
		#head_color = (0.8,0.8,0.8,1)	# Цвет фона заголовка таблицы
		table_color = (0.8,0.98,0.98,1)
		col_range, row_range = Excel.get_cols_rows_names(data_range)
		cols_width = {pos: mm(sheet.column_dimensions[col].width) for pos, col in enumerate(col_range)}
		rows_height = {pos: mm(sheet.row_dimensions[row].height) for pos, row in enumerate(row_range)}
		widget.columns_width = cols_width
		# Записываем данные в таблицу
		#head_data = [c.value for r in sheet[headers_range] for c in r]
		table_data = [c.value for r in sheet[data_range] for c in r]
		widget.data = [{'text': str(data) if data!=None else "", 'background_color': table_color, 'focus': False} for data in table_data]

# class ScladItem():
# 	type = ''
# 	mount_style = ''
# 	code = ''
# 	order_code = ''
# 	analog = ''
# 	count = 0
# 	description = ''
# 	comment = ''
# 	manufacturer = ''
# 	price = ''
# 	date = ''


def create_sclad(main_table='', devs=[], specs={}):
	""" Берет списки комплектующих устройств и добавляет позиции на склад, 
	если они туда еще не добавлены
	main_table - путь к файлу склада в формате 'folder\path\file_name.xlsx:sheet_name'
	devs - список устройств, со спецификаций которых добавляются позиции на склад
	specs - словарь оссоциации устройства и файла со спецификацией формата:
	specs = {'dev1': 'folder\path\file_name.xlsx:sheet_name'}"""
	# #import pandas
	# from openpyxl.utils.dataframe import dataframe_to_rows
	# import numpy
	# t1 = pandas.DataFrame({'A':[0, 1, 2, 3], 'B':['a', 'b', 'c', 'd'], 
	# 	'C':['0', '1', '2', '3']})
	# print(t1)
	# print('-----------')
	# t2 = pandas.DataFrame({'A':[4, 5, 2, 3, 6, 7], 'B':['a', 'b', 'c', 'd', 'e', 'f'], 
	# 	'D':[12, 23, 34, 45, 56, 67]})
	# print(t2)
	# print('-----------')
	# repl = {'A1': pandas.Series(['a', 'A', 'AA']), 
	# 		'B1': pandas.Series(['b', 'B'])}
	# rt = pandas.DataFrame(repl)
	# print(rt)
	#t2.columns = t2.columns[rt.isin(t2.columns)]
	#print(t2.columns)
	# # print(t2['A'][t2['A']>5])
	# # Получаем список ненужных столбцов
	# ic = ['A', 'B', 'C']
	# print(t2.columns)
	# dc = t2.columns[~t2.columns.isin(ic)]
	# print(dc)
	# # Удаляем ненужные столбцы
	# t2 = t2.drop(columns=dc)
	# print(t2)
	# print("--------")
	# t3 = t2[~t2['A'].isin(t1['A'])]
	# print(t3)
	# print('----------')
	# t4 = t1.append(t3, ignore_index=True)
	# print(t4)
	# wb = Workbook()
	# ws = wb.active
	# for r in dataframe_to_rows(t4, index=False, header=True):
	# 	ws.append(r)
	# print(ws.values)
	# wb.save('res.xlsx')
	if not main_table or not devs or not specs:
		print("[ERROR]: Нет входных данных!")
		return
	table_xl, table_sh = main_table.split(':')
	#print(table_xl, table_sh)
	sclad_wb = Excel.open_book(table_xl)
	sclad_sh = sclad_wb[table_sh]
	df = pandas.DataFrame(sclad_sh.values)
	sclad = get_spec(df)
	#print(sclad['Order Code'])
	print(type(sclad))
	if len(sclad) == 0:
		print(f"[Error]: Ошибка обработки {sclad_sh}!")
		return
	# Список с названиями интересующих нас столбцов:
	interest_cols = sclad.columns
	#interest_cols = ['Type', 'MountStyle', 'Code', 'Order Code', 'Analog', 'Count', 
	#	'Price', 'Date', 'Manufacturer', 'Package', 'Description', 'Comment']
	replacement_names = { 'Order Code': ['OrderCode', 'Order_Code'],
		'Type': ['Класс', 'Class', 'class', 'класс'], 
		'MountStyle': ['Column=MountStyle', 'Mount Style'],
		'Description': ['Column=Description'],
		'Comment': ['Column=Comment'],
		'Supplier': ['Supplier 1']}
	unprocessed = []
	fails = 0
	pas = 0
	for dev in devs:
		path = specs.get(dev, 'None')
		if path != 'None':
			xl, sh = path.split(':')
			sheet = Excel.open_book(xl)[sh]
			df = pandas.DataFrame(sheet.values)
			tbl = get_spec(df, replacement_names)	# Получаем таблицу спецификации
			# Проверяем валидность таблицы
			if len(tbl) == 0:
				print(f"[Error]: {dev} не обработан!")
				fails += 1
				continue
			print(f"[PASS]: {dev} обработан")
			# Получаем таблицу с позициями, которые отсутствуют в таблице склада
			merge_tbl = tbl[~tbl['Order Code'].isin(sclad['Order Code'])]
			# Получаем список ненужных столбцов
			delcol = merge_tbl.columns[~merge_tbl.columns.isin(interest_cols)]
			# Удаляем ненужные столбцы
			merge_tbl = merge_tbl.drop(columns=delcol)
			# Сливаем таблицы, без индексов, с названиями столбцов
			sclad = sclad.append(merge_tbl, ignore_index=True)
			pas +=1
		else: 
			unprocessed.append(dev)
			fails += 1
			print(f"[Error]: {dev} не обработан!")
	print(f"Обработка завершена: PASS: {pas}; FAILS: {fails}; Всего: {len(devs)})")
	# Сохраняем резултат
	#wb = Excel.open_book('res.xlsx')
	#ws = wb['Sheet']
	#row_dim = ws.row_dimensions
	Excel.clear_sheet(sclad_sh) #, del_formatting=False)
	for r in dataframe_to_rows(sclad, index=False, header=True):
		sclad_sh.append(r)
	#ws.row_dimensions = row_dim
	#sclad_wb.save('res.xlsx')
	sclad_wb.save(table_xl)
	print("Таблица сохранена.")

def get_spec(table, replacement_names={}):
	""" Находит и возвращает таблицу спецификации.  
	Индексирует названия столбцов из таблицы, чтобы по ним было проще выбирать данные.
	table - таблица в формате DataFrame
	replacement_names - ключи для замены имен столбцов с неправильных на правильные:
	replacement_names = {pass_name: [fail_Name1, fail_name2, ...]}"""
	#print(table.shape)	# Получить число строк и колонок
	tbl = pandas.DataFrame()
	for r, row in table.iterrows():
		if 'Order Code' in row.values:
			# Находим строку с названиями столбцов
			table.columns = list(row)	# Присваиваем названия столбцов из таблицы
			tbl = table.iloc[r+1:]	# Отсекаем первые строки с ненужной информацией
			break
	# Форматируем словарь соответствия имен столбцов таблицы для переименования
	repl_name = {}
	for pn in replacement_names.keys():
		repl_name.update(dict.fromkeys(replacement_names[pn], pn))
	# Переименовываем столбцы на правильные названия по словарю соответствия
	tbl = tbl.rename(columns = repl_name)
	return tbl

# def create_parts_list(devs=[], specs={}):
# 	""" Создает общий список комплектующих для устройств devs.
# 	Спецификация берется из словаря specs, где ключ - название изделя, значение - таблица спецификации
# 	devs - list type
# 	specs - dict type"""
# 	unaccounted = []
# 	for dev in devs:
# 		sp = specs.get(dev)
# 		if sp:
# 			print(sp)
# 		else: 
# 			unaccounted.append(dev)
# 	print(unaccounted)

def podschet():
	"""
	Скрипт подсчитывает общее количество компонентов для нескольких проектов. Если в разных проектах находятся
одинаковые позиции - количество складывается.
Необходимо доработать алгоритм - Если позиция в одном изделии называется например "ADUM1201D" и указан 
его аналог "ADUM1201D-REEL7", а во втором изделии указана позиция "ADUM1201D-REEL7" - то позиции не сольются.тСкрипт не ищет по аналогам, он не найдет позицию "ADUM1201D-REEL7" в уже имеющихся и добавит ее как новую 
позицию.
	Так же необходимо убедиться, чтобы не учитывать регистр.
	"""
	def sort(vals, kols, analog):
		# Функция берет значения из текущего списка, сравнивает и добавляет их к общему списку с учетом количества
		# Если значение уже есть в общем списке - прибавляем к нему количество.
		# Если нет - добавляем новую позицию.
		identical = 0
		added = 0
		i = 0
		for i, item in enumerate(vals):
		# i - текущий индекс цикла; item - текущее значение по индексу в цикле
			#print(f"item = {i}: {item}")
			if item in vals_new:
				print(f"Идентично: {vals_new.index(item)}:{vals_new[vals_new.index(item)]}={kols_new[vals_new.index(item)]} + {i}:{item}={kols[i]}")
				#прибавляем количество по item к количеству по vals_new.index(item)
				kols_new[vals_new.index(item)] += kols[i]
				#записываем аналог 
				print(f"an_new: {analog_new[vals_new.index(item)]}")
				print(f"an: {analog[i]}")
				if analog_new[vals_new.index(item)] != analog[i]:
					if analog[i] != None:
						if analog_new[vals_new.index(item)] != None:
							analog_new[vals_new.index(item)] += "; "
						else:
							analog_new[vals_new.index(item)] = ""
						analog_new[vals_new.index(item)] += analog[i]
				identical += 1
			else:
				vals_new.append(item)   #добавили элемент в новый список
				kols_new.append(kols[i]) #добавили количество
				analog_new.append(analog[i])    #добавили аналог
				added += 1
				#print(" (Добавлено)")
		print(f"Обработка таблицы завершена. Идентичных = {identical}; Добавлено = {added}")

	print("Начинаем работу с файлом...")
	#создать COM объект
	#Excel = win32com.client.Dispatch("Excel.Application")
	#Excel = client.Dispatch("Excel.Application")
	
	# Открыть файл Excel для работы
	#wb = Excel.open_book(u'D:\\LMT\\SCLAD\\Закупка февраль 2021\\Общий.xlsx')
	wb = Excel.open_book(u'Общий.xlsx')

	#Определяем нужные листы
	#sh_res = wb.Worksheets("Израсходовано")
	sh_res = wb['Итог']
	sh_spr4 = wb["SPR4-03"]
	sh_sim868 = wb["SOVA_R1.1-04"]
	sh_pcf8253 = wb["BUZINA R1 - 01"]
	sh_tavolga = wb["KERRIA_R1.1-01"]
	sh_spmd = wb["MCN3_R3-01"]
	sh_lbox = wb["TAVOLGA4U R2.1-01"]
	sh_lbox_ext = wb["SPMD1702M1_R2-01"]

	#Очищаем таблицу результата
	##sh_res.Cells.Clear
	#sh_res.delete_rows(0, sh_res.max_row)	# Удаляет все полностью.
	Excel.clear_sheet(sh_res)

	print("ROWS IN SHEET:", sh_res.max_row)
	#wb.save(u'D:\\LMT\\SCLAD\\Закупка февраль 2021\\Общий.xlsx')

	#начинаем работать с данными

	#получить значение одной ячейки
	#val = sh_sim868.Cells(3,5).value
	#print(f"Прочитано: {val}")

	#получить значения диапазона ячеек
	vals_new = []       #пустой список OrderCode для записи результата
	kols_new = []        #пустой список количество для записи результата
	analog_new = []         #пустой список Analog для записи результата

	vals_spr4 = [r[0].value for r in sh_spr4["F4:F32"]]       # OrderCode
	kols_spr4 = [r[0].value for r in sh_spr4["I4:I32"]]       # количество
	an_spr4 = [r[0].value for r in sh_spr4["G4:G32"]]       # Analog

	vals_sim868 = [r[0].value for r in sh_sim868["F4:F32"]]       # OrderCode
	kols_sim868 = [r[0].value for r in sh_sim868["J4:J32"]]       # количество
	an_sim868 = [r[0].value for r in sh_sim868["G4:G32"]]       # Analog

	vals_pcf8253 = [r[0].value for r in sh_pcf8253["F4:F56"]]       # OrderCode
	kols_pcf8253 = [r[0].value for r in sh_pcf8253["J4:J56"]]       # количество
	an_pcf8253 = [r[0].value for r in sh_pcf8253["G4:G56"]]       # Analog

	vals_tavolga = [r[0].value for r in sh_tavolga["F4:F71"]]     # OrderCode
	kols_tavolga = [r[0].value for r in sh_tavolga["J4:J71"]]     # количество
	an_tavolga = [r[0].value for r in sh_tavolga["G4:G71"]]       # Analog

	vals_spmd = [r[0].value for r in sh_spmd["F4:F86"]]           # OrderCode
	kols_spmd = [r[0].value for r in sh_spmd["J4:J86"]]           # количество
	an_spmd = [r[0].value for r in sh_spmd["G4:G86"]]       # Analog

	vals_lbox = [r[0].value for r in sh_lbox["C4:C86"]]   # OrderCode
	kols_lbox = [r[0].value for r in sh_lbox["F4:F86"]]   # количество
	an_lbox = [r[0].value for r in sh_lbox["D4:D86"]]       # Analog

	vals_lbox_ext = [r[0].value for r in sh_lbox_ext["F4:F78"]]    # OrderCode
	kols_lbox_ext = [r[0].value for r in sh_lbox_ext["I4:I78"]]    # количество
	an_lbox_ext = [r[0].value for r in sh_lbox_ext["G4:G78"]]       # Analog

	print("Данные с таблиц прочитаны.")

	#формируем новый список комплектующих с подсчетом количества
	print(an_spr4)
	sort(vals_spr4, kols_spr4, an_spr4)
	print(analog_new)
	print(an_sim868)
	sort(vals_sim868, kols_sim868, an_sim868)
	print(analog_new)
	sort(vals_pcf8253, kols_pcf8253, an_pcf8253)
	sort(vals_tavolga, kols_tavolga, an_tavolga)
	sort(vals_spmd, kols_spmd, an_spmd)
	sort(vals_lbox, kols_lbox, an_lbox)
	sort(vals_lbox_ext, kols_lbox_ext, an_lbox_ext)

	#Выписываем вначало листа наименование и количество плат, ушедшее в расчет
	#Записать значения одной ячейки
	sh_res.cell(row=1, column=1).value = sh_spr4.title
	sh_res.cell(row=1, column=3).value = sh_spr4.cell(row=1, column=2).value
	sh_res.cell(row=2, column=1).value = sh_sim868.title
	sh_res.cell(row=2, column=3).value = sh_sim868.cell(row=1, column=2).value
	sh_res.cell(row=3, column=1).value = sh_pcf8253.title
	sh_res.cell(row=3, column=3).value = sh_pcf8253.cell(row=1, column=2).value
	sh_res.cell(row=4, column=1).value = sh_tavolga.title
	sh_res.cell(row=4, column=3).value = sh_tavolga.cell(row=1, column=2).value
	sh_res.cell(row=5, column=1).value = sh_spmd.title
	sh_res.cell(row=5, column=3).value = sh_spmd.cell(row=1, column=2).value
	sh_res.cell(row=6, column=1).value = sh_lbox.title
	sh_res.cell(row=6, column=3).value = sh_lbox.cell(row=1, column=2).value
	sh_res.cell(row=7, column=1).value = sh_lbox_ext.title
	sh_res.cell(row=7, column=3).value = sh_lbox.cell(row=1, column=2).value

	Excel.create_template(sh_res)	# Загружаем шаблон странцы

	#sh_res.cell(row=9, column=1).value = "OrderCode"
	#sh_res.cell(row=9, column=2).value = "Analog"
	#sh_res.cell(row=9, column=3).value = "Необходимо"

	#записать значения диапазона
	align = Alignment(horizontal="center", wrap_text=True)	# Выравнивание по центру и перенос текста
	r = sh_res.max_row + 1
	i = r
	for rec in vals_new:
		sh_res.cell(i,1).value = rec
		sh_res.cell(i,1).alignment = align
		i += 1
	i = r
	for an in analog_new:
		sh_res.cell(i,2).value = an
		sh_res.cell(i,2).alignment = align
		i += 1
	i = r
	for kol in kols_new:
		sh_res.cell(i,3).value = kol
		i += 1

	#sh_res.calculate_dimension()
	#sh_res.row_dimensions.update()
	#sh_res.column_dimensions.update()
	Excel.save_book(wb, u'Общий.xlsx')
	wb.close()
	#print(f"Данные записаны в лист: {sh_res.title}")


def prihod_compel():
	"""
Модуль подсчитывает закупленные комплектующие в Компэле и записывает их в общий файл "Приход"
Источник: .xls таблицы из Компэла формата "orders_web_". Считываются все файлы этого типа в указанном каталоге.
Результат: Список закупленных комплектующих в файле "Приход.xls". 
Если попадаются 2 одинаковых позиции - их количество суммируется.
"""
	def sort(vals, kols):
		# Функция берет значения из текущего списка, сравнивает и добавляет их к общему списку с учетом количества
		# Если значение уже есть в общем списке - прибавляем к нему количество.
		# Если нет - добавляем новую позицию.
		identical = 0
		added = 0
		i = 0
		for i, item in enumerate(vals):
		# i - текущий индекс цикла; item - текущее значение по индексу в цикле
			#print(f"item = {i}: {item}")
			if item in vals_new:
				#прибавляем количество по item к количеству по vals_new.index(item)
				print(f"Идентично: {vals_new.index(item)}:{vals_new[vals_new.index(item)]}={kols_new[vals_new.index(item)]} + {i}:{item}={kols[i]}")
				kols_new[vals_new.index(item)] += kols[i]
				identical += 1
			else:
				vals_new.append(item)   #добавили элемент в новый список
				kols_new.append(kols[i]) #добавили количество
				added += 1
				#print(" (Добавлено)")
		print(f"Обработка таблицы завершена. Идентичных = {identical}; Добавлено = {added}")

	print("Начинаем работу с файлом...")
	#создать COM объект
	Excel = win32com.client.Dispatch("Excel.Application")
	#Excel = client.Dispatch("Excel.Application")
	#открыть файл Excel для работы
	wb = Excel.Workbooks.Open(u'D:\\SCLAD\\Закупка октябрь 2020\\Приход.xlsx')

	#Обработаем каталог, где лежат счета
	directory = 'D:\\SCLAD\\Закупка октябрь 2020\\счета'
	all_files = os.listdir(directory)   		#получить список всех файлов в каталоге
	xl_files = [f for f in os.listdir(directory) if f.endswith('.xls')]  #отфильтровать файлы Excel
	#order_files = [f for f in os.listdir(directory) if f.startswith('Зак_')] #отфильтровать файлы счетов
	#opis_files = [f for f in os.listdir(directory) if f.startswith('Opis_gruza')] #отфильтровать файлы "Опись груза"
	print(all_files)
	print(xl_files)

	#Определяем нужные листы
	sh_res = wb.Worksheets("Итог")
	#Очищаем таблицу результата
	sh_res.Cells.Clear()
	sh_res.Range("A1").Select()
	wb.Save()   #сохраняем таблицу чтобы убрать старые границы

	#начинаем работать с данными
	vals_new = []       #пустой список OrderCode для записи результата
	kols_new = []        #пустой список количество для записи результата

	for curfile in xl_files:
		xl = Excel.Workbooks.Open(directory + "\\" + curfile)  #Открыли порядковый файл котолога
		#Вычитываем значения из файла
		lastrow = xl.ActiveSheet.UsedRange.Rows.Count - 1   #определили последнюю ячейку файла
		#print(f"{curfile}: {lastrow}")
		vals_sheet = [r[0].value for r in xl.ActiveSheet.Range(f"E6:E{lastrow}")]
		kols_sheet = [r[0].value for r in xl.ActiveSheet.Range(f"I6:I{lastrow}")]
		sort(vals_sheet, kols_sheet)
	xl.Close()

	#получить значение одной ячейки
	#val = sh_sim868.Cells(3,5).value
	#получить значения диапазона ячеек
	#vals_spr4 = [r[0].value for r in sh_spr4.Range("F4:F31")]       # OrderCode
	#kols_spr4 = [r[0].value for r in sh_spr4.Range("I4:I31")]       # количество

	#Записать значения одной ячейки
	sh_res.Cells(1,1).value = "OrderCode"
	sh_res.Cells(1,2).value = "Quantity"
	#sh_res.Cells(9,3).value = "Необходимо"

	#записать значения диапазона
	i = 2
	for rec in vals_new:
		sh_res.Cells(i,1).value = rec
		i += 1
	i = 2
	for kol in kols_new:
		sh_res.Cells(i,2).value = kol
		i += 1

	print(f"Данные записаны в лист: {sh_res.Name}")

	wb.Save()
	wb.Close()
	Excel.Quit()


def sortirovka_prihod():
	"""
	Модуль сортирует закупленные комплектующие из файла "Приход.xls" в файл "Общий.xls".
Поиск позиций происходит по графам OrderCode и Analog
При записи отсортированных данных в графу Закуплено старые данные суммируются или перезаписываются
в зависимости от флага fl_sum_kols.
	"""
	#Флаг указывает суммировать ли новые данные со старыми в графе закуплено. 
	#True - суммировать количество, False - перезаписать
	fl_sum_kols = True

	def sort(res_vals, res_analog, old_vals, old_kols, zak_vals, zak_kols):
		# Функция берет значения из закупочного списка, сравнивает с результирующем списком (OrderCode и Analog)
		# и формирует новый список в зависимости от результата
		# Если совпадений с OrderCode или Analog не нашлось - формируется другой список "Остаток"
		#global vals_new, kols_new, vals_ost, kols_ost, fl_sum_kols
		identical = 0
		added = 0
		print("Обрабатываем по графе OrderCode...")
		i = 0
		for i, item in enumerate(res_vals):
			#проходим все позиции OrderCode для заказа
			# i - текущий индекс цикла; item - текущее значение по индексу в цикле
			if item in zak_vals:
				#добавляем соответствующий zak_vals и zak_kols к текущей позиции res_vals
				print(f"{item} ({zak_kols[zak_vals.index(item)]}) - Есть в zak_vals")
				#vals_new.append(item)
				if fl_sum_kols:
					if old_vals[i] == None:
						old_vals[i] = ""
					if old_vals[i] != "":
						old_vals[i] = str(old_vals[i]) + "; "
					vals_new.append(str(old_vals[i]) + str(item))
					kols_new.append(old_kols[i] + zak_kols[zak_vals.index(item)])
				else:
					vals_new.append(item)
					kols_new.append(zak_kols[zak_vals.index(item)])
				#удаляем соответствующий zak_vals и zak_kols
				zak_kols.pop(zak_vals.index(item))
				zak_vals.pop(zak_vals.index(item))
			else:
				print(f"{item} - Нет в zak_vals")
				#добавляем пустые значения zak_vals и zak_kols к текущей позиции res_vals
				#vals_new.append("")
				if fl_sum_kols:
					if old_vals[i] == None:
						old_vals[i] = ""
					vals_new.append(old_vals[i])
					kols_new.append(old_kols[i])
				else:
					vals_new.append("")
					kols_new.append(0)
		print("Обрабатываем по графе Analog...")
		i = 0
		for i, item in enumerate(res_analog):
			#проходим все позиции Analog для заказа
			#обработать item - разложить если несколько значений
			s = 0
			for s, spl in enumerate(str(item).split('; ')):
				if spl in zak_vals:
					#добавляем (+=) соответствующий zak_vals и zak_kols к текущей позиции res_vals
					print(f"{item}: {spl} - Есть в zak_vals")
					if vals_new[i] != "":
						vals_new[i] += "; "
					vals_new[i] += spl
					kols_new[i] += zak_kols[zak_vals.index(spl)]
					#удаляем соответствующий zak_vals и zak_kols
					zak_kols.pop(zak_vals.index(spl))
					zak_vals.pop(zak_vals.index(spl))
				else:
					print(f"{item}: {spl} - Нет в zak_vals")

		vals_ost = zak_vals
		kols_ost = zak_kols
		print(f"Обработка таблицы завершена. Идентичных = {identical}; Добавлено = {added}")

	print("Начинаем работу с файлом...")
	#создать COM объект
	Excel = win32com.client.Dispatch("Excel.Application")
	#Excel = client.Dispatch("Excel.Application")
	#открыть файл Excel для работы
	wb = Excel.Workbooks.Open(u'D:\\SCLAD\\Закупка октябрь 2020\\Общий.xlsx')
	wb_zak = Excel.Workbooks.Open(u'D:\\SCLAD\\Закупка октябрь 2020\\Приход.xlsx')

	#Определяем нужные листы
	#sh_res = wb.Worksheets("Израсходовано")
	sh_res = wb.Worksheets("Итог")
	sh_ost = wb.Worksheets("Неучтено2")
	sh_zak = wb_zak.Worksheets("Итог")

	#Очищаем таблицу результата
	#sh_res.Cells.Clear

	#начинаем работать с данными

	#получить значение одной ячейки
	#val = sh_sim868.Cells(3,5).value
	#print(f"Прочитано: {val}")

	#получить значения диапазона ячеек
	vals_new = []       #пустой список OrderCode для записи результата
	kols_new = []        #пустой список количество для записи результата
	vals_ost = []       #пустой список OrderCode для записи остатка
	kols_ost = []       #пустой список количество для записи остатка

	lastrow_zak = sh_zak.UsedRange.Rows.Count   #определили последнюю ячейку файла
	lastrow_res = sh_res.UsedRange.Rows.Count   #определили последнюю ячейку файла

	vals_res = [r[0].value for r in sh_res.Range(f"A10:A{lastrow_res}")]
	analog_res = [r[0].value for r in sh_res.Range(f"B10:B{lastrow_res}")]
	vals_old = [r[0].value for r in sh_res.Range(f"E10:E{lastrow_res}")]
	kols_old = [r[0].value for r in sh_res.Range(f"F10:F{lastrow_res}")]

	vals_zak = [r[0].value for r in sh_zak.Range(f"A2:A{lastrow_zak}")]
	kols_zak = [r[0].value for r in sh_zak.Range(f"B2:B{lastrow_zak}")]

	print("Данные с таблиц прочитаны.")

	#формируем новый список комплектующих с учетом аналогов
	sort(vals_res, analog_res, vals_old, kols_old, vals_zak, kols_zak)

	#Записать значения одной ячейки
	sh_res.Cells(9,5).value = "Позиция"
	sh_res.Cells(9,6).value = "Закуплено"

	#записать значения диапазона
	i = 10
	for rec in vals_new:
		sh_res.Cells(i,5).value = rec
	i += 1
	i = 10
	for kol in kols_new:
		sh_res.Cells(i,6).value = kol
		i += 1

	print(f"Данные записаны в лист: {sh_res.Name}")

	#Записать остаток
	sh_ost.Cells(1,1).value = "Позиция"
	sh_ost.Cells(1,2).value = "Закуплено"

	#Очищаем таблицу результата
	sh_ost.Cells.Clear

	i = 2
	for rec in vals_ost:
		sh_ost.Cells(i,1).value = rec
		i += 1
	i = 2
	for kol in kols_ost:
		sh_ost.Cells(i,2).value = kol
		i += 1
	print(f"Данные записаны в лист: {sh_ost.Name}")

	wb.Save()
	wb.Close()
	Excel.Quit()


def sortirovka_rashod():
	def sort(res_vals, res_analog, res_kols, rash_vals, rash_kols):
	# Функция берет значения из закупочного списка, сравнивает с результирующем списком (OrderCode и Analog)
	# и формирует новый список в зависимости от результата
	# Если совпадений с OrderCode или Analog не нашлось - формируется другой список "Остаток"
		identical = 0
		added = 0
		print("Обрабатываем по графе OrderCode...")
		i = 0
		for i, item in enumerate(res_vals):
			#проходим все позиции OrderCode для заказа
			# i - текущий индекс цикла; item - текущее значение по индексу в цикле
			if item in rash_vals:
				#добавляем соответствующий zak_vals и zak_kols к текущей позиции res_vals
				print(f"{item} ({rash_kols[rash_vals.index(item)]}) - Есть в rash_vals")
				vals_new.append(item)
				kols_new.append(rash_kols[rash_vals.index(item)])
				#удаляем соответствующий zak_vals и zak_kols
				#zak_kols.pop(zak_vals.index(item))
				#zak_vals.pop(zak_vals.index(item))
			else:
				print(f"{item} - Нет в zak_vals")
				#добавляем пустые значения zak_vals и zak_kols к текущей позиции res_vals
				vals_new.append("")
				kols_new.append(0)
		print("Обрабатываем по графе Analog...")
		"""
		i = 0
		for i, item in enumerate(res_analog):
			#проходим все позиции Analog для заказа
			#обработать item - разложить если несколько значений
			#s = 0
			for s, spl in enumerate(str(item).split('; ')):
				if spl in zak_vals:
					#добавляем (+=) соответствующий zak_vals и zak_kols к текущей позиции res_vals
					print(f"{item}: {spl} - Есть в zak_vals")
					if vals_new[i] != "":
						vals_new[i] += "; "
					vals_new[i] += spl
					kols_new[i] += zak_kols[zak_vals.index(spl)]
					#удаляем соответствующий zak_vals и zak_kols
					zak_kols.pop(zak_vals.index(spl))
					zak_vals.pop(zak_vals.index(spl))
				else:
					print(f"{item}: {spl} - Нет в zak_vals")
		"""
		print(f"Обработка таблицы завершена. Идентичных = {identical}; Добавлено = {added}")

	print("Начинаем работу с файлом...")
	#создать COM объект
	Excel = win32com.client.Dispatch("Excel.Application")
	#Excel = client.Dispatch("Excel.Application")
	#открыть файл Excel для работы
	wb = Excel.Workbooks.Open(u'D:\\SCLAD\\Закупка август 2020\\Общий.xlsx')

	#Определяем нужные листы
	#sh_res = wb.Worksheets("Израсходовано")
	sh_res = wb.Worksheets("Итог")
	sh_rash = wb.Worksheets("Расход")

	#Очищаем таблицу результата
	#sh_res.Cells.Clear

	#начинаем работать с данными

	#получить значение одной ячейки
	#val = sh_sim868.Cells(3,5).value
	#print(f"Прочитано: {val}")

	#получить значения диапазона ячеек
	vals_new = []       #пустой список OrderCode для записи результата
	kols_new = []        #пустой список количество для записи результата

	lastrow_res = sh_res.UsedRange.Rows.Count   #определили последнюю ячейку файла
	lastrow_rash = sh_rash.UsedRange.Rows.Count   #определили последнюю ячейку файла

	vals_res = [r[0].value for r in sh_res.Range(f"A10:A{lastrow_res}")]
	analog_res = [r[0].value for r in sh_res.Range(f"B10:B{lastrow_res}")]
	kols_res = [r[0].value for r in sh_res.Range(f"C10:C{lastrow_res}")]

	vals_rash = [r[0].value for r in sh_rash.Range(f"A10:A{lastrow_rash}")]
	kols_rash = [r[0].value for r in sh_rash.Range(f"C10:C{lastrow_rash}")]

	print("Данные с таблиц прочитаны.")

	#формируем новый список комплектующих с учетом аналогов
	sort(vals_res, analog_res, kols_res, vals_rash, kols_rash)

	#Записать значения одной ячейки
	sh_res.Cells(9,7).value = "Израсходованно"

	#записать значения диапазона
	"""
	i = 10
	for rec in vals_new:
		sh_res.Cells(i,5).value = rec
		i += 1
	"""
	i = 10
	for kol in kols_new:
		sh_res.Cells(i,7).value = kol
		i += 1

	#Записать остаток

	print(f"Данные записаны в лист: {sh_res.Name}")

	wb.Save()
	wb.Close()
	Excel.Quit()
	


def sortirovka_v_zacupku():
	def sort(vals, analog, kols):
		added = 0
		print("Обрабатываем по графе OrderCode...")
		i = 0
		for i, item in enumerate(kols):
			if item > 0:
				vals_new.append(vals[i])
				analog_new.append(analog[i])
				kols_new.append(item)
				added += 1

		print(f"Обработка таблицы завершена. Добавлено = {added}")

	print("Начинаем работу с файлом...")
	#создать COM объект
	Excel = win32com.client.Dispatch("Excel.Application")
	#Excel = client.Dispatch("Excel.Application")
	#открыть файл Excel для работы
	wb = Excel.Workbooks.Open(u'D:\\SCLAD\\Закупка август 2020\\Общий.xlsx')
	wb1 = Excel.Workbooks.Open(u'D:\\SCLAD\\Закупка август 2020\\В закупку.xlsx')

	#Определяем нужные листы
	#sh_res = wb.Worksheets("Израсходовано")
	sh_ish = wb.Worksheets("Итог")
	sh_zak = wb1.Worksheets("В закупку")

	#Очищаем таблицу результата
	sh_zak.Cells.Clear

	#начинаем работать с данными

	#получить значение одной ячейки
	#val = sh_sim868.Cells(3,5).value
	#print(f"Прочитано: {val}")

	#получить значения диапазона ячеек
	vals_new = []       #пустой список OrderCode для записи результата
	kols_new = []        #пустой список количество для записи результата
	analog_new = []

	lastrow_ish = sh_ish.UsedRange.Rows.Count   #определили последнюю ячейку файла
	#lastrow_rash = sh_rash.UsedRange.Rows.Count   #определили последнюю ячейку файла

	vals_ish = [r[0].value for r in sh_ish.Range(f"A10:A{lastrow_ish}")]
	analog_ish = [r[0].value for r in sh_ish.Range(f"B10:B{lastrow_ish}")]
	kols_ish = [r[0].value for r in sh_ish.Range(f"J10:J{lastrow_ish}")]

	#vals_rash = [r[0].value for r in sh_rash.Range(f"A10:A{lastrow_rash}")]
	#kols_rash = [r[0].value for r in sh_rash.Range(f"C10:C{lastrow_rash}")]

	print("Данные с таблиц прочитаны.")

	#формируем новый список комплектующих с учетом аналогов
	sort(vals_ish, analog_ish, kols_ish)

	#Записать значения одной ячейки
	sh_zak.Cells(1,1).value = "Наименование"
	sh_zak.Cells(1,2).value = "Аналог"
	sh_zak.Cells(1,3).value = "Количество"

	#записать значения диапазона
	i = 2
	for rec in vals_new:
		sh_zak.Cells(i,1).value = rec
		i += 1

	i = 2
	for an in analog_new:
		sh_zak.Cells(i,2).value = an
		i += 1

	i = 2
	for kol in kols_new:
		sh_zak.Cells(i,3).value = kol
		i += 1

	print(f"Данные записаны в лист: {sh_zak.Name}")

	wb.Save()
	wb.Close()
	wb1.Save()
	wb1.Close()
	Excel.Quit()

if __name__ == '__main__':
	import configparser
	config = configparser.ConfigParser()
	config.optionxform = str	# Отключаем анализ файлов Windows INI, чтобы не менялся регистр для ключей
	config.read("bindings.ini")
	main_table = config['files']['main_xls']
	#main_table_xl, main_table_sh = main_table.split(':')
	#print(main_table)
	#print(main_table_xl, main_table_sh)
	devs_list_str = config['general']['devs_list']
	devs_list = devs_list_str.split(', ')	#[type: list] Получили список устройств для обработки
	#print(devs_list)
	binders = dict(config.items('binders'))
	#print(binders)
	create_sclad(main_table = main_table, devs = devs_list, specs = binders)


	#from openpyxl.utils.cell import cols_from_range
	#from openpyxl.utils.cell import column_index_from_string
	#podschet()
	#wb = Excel.open_book(u'Общий.xlsx')
	#ws = wb['Итог'] #['SPR4-03']
	# Одна единица ширины столбца равна ширине одного символа в стиле Normal. 
	# Для пропорциональных шрифтов используется ширина символа 0 (ноль).
	# Вычитанная ширина столбца 'B' = 22.7109375. В Excel показывает ширину 22 едениц и 159 пикселя (в пикселях верно)
	# Если прочитать ячейку: ws.column_dimensions['B1'].width = 13
	#  ws.column_dimensions['B'].font.sz = 11.0	(Для колонки 'A' тоже 11)
	#cd = {0: 184, 1: 159, 2: 123, 3: 94, 4: 143, 5: 80, 6: 116, 7: 79, 8: 64, 9: 71} # Правильные значения в пикселях
	#print(wb.loaded_theme)
	#print(dir(ws))
	#cd0 = {Excel._ord_col_name(k[0])-1: k[1].width for k in ws.column_dimensions.items()}
	#print("items =", cd0)
	#print("width =", ws.column_dimensions['B'].width)
	#print("font", ws.column_dimensions['B'].font)
	# ws.conditional_formatting
	#print("other", ws.column_dimensions['B'].font)
	#print("+++++++++++")
	#cr = [col for col in cols_from_range("A1:F5")]
	#cr = column_index_from_string('A')
	#print("column_range:", cr)
